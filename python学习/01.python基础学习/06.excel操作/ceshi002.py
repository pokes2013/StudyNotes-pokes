from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active

# 修改a6
ws["a6"] = "冰冷的希望"

wb.save("test.xlsx")