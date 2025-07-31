# 导入模块
import openpyxl

workbook = openpyxl.load_workbook("example.xlsx")

# 获取活动表对象
sheet = workbook.active
# 通过行和列得到指定单元格对象
cell1 = sheet.cell(row=1,column=1)      # 表示获得第一行第一列的单元格对象
cell2 = sheet.cell(row=2,column=2)      # 表示获得第2行第3列的单元格对象
print(cell1,cell2)      # <Cell 'Sheet1'.A1> <Cell 'Sheet1'.B2> ，代表A1和C2两个单元格对象

# 通过单元格对象得到单元格中的数据
res1 = cell1.value      # 获得A1中的数据
res2 = cell2.value      # 获得C2中的数据
print(res1,res2)        # 输出结果：姓名 25
