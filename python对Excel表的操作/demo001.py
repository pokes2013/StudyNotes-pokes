from openpyxl import Workbook, load_workbook

# 创建新 Excel 文件
wb = Workbook()     #创建一个工作表对象
ws = wb.active      #激活这个对象
ws.title = "Sheet1"     #表的名称

# 写入数据
ws['A1'] = "姓名"
ws['B1'] = "年龄"
ws['A2'] = "张三"
ws['B2'] = 25
ws['A3'] = "李四"
ws['B3'] = 24


# 保存文件
wb.save("example.xlsx")

# 读取 Excel 文件
wb = load_workbook("example.xlsx")
ws = wb.active

# 读取数据
##读取每一行的数据，每一行的数据就是一个列表，使用逗号分隔；
for row in ws.iter_rows(values_only=True):
    print(row)
